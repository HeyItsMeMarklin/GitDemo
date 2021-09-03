import openpyxl


class HomePageData:

    test_home_page_data = [{"first_name": "Marklin", "email": "viraymarklin12@gmail.com", "password": "testMarklin123#",
                            "gender": "Male"},
                           {"first_name": "Ellie", "email": "elliemaewong1998@gmail.com", "password": "ellielin122326",
                            "gender": "Female"}]

    @staticmethod
    def get_data_from_excel(test_case_name):
        data_dictionary = {}
        book = openpyxl.load_workbook("C:\\Users\\Marklin\\Desktop\\python_excel_demo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    data_dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [data_dictionary]
