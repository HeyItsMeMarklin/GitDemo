import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Marklin\\Desktop\\python_excel_demo.xlsx")

sheet = book.active
data_dictionary = {}
cell = sheet.cell(row=2, column=2)
print(cell.value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=1,column=1).value == "test_case_2":
        for j in range(1, sheet.max_column+1):
            data_dictionary[sheet.cell(row=i, column=j)] = sheet.cell(row=i, column=j).value

print(data_dictionary)